"""
Econet Editora — Scraper PIS/COFINS por NCM
============================================
- 1ª execução: abre browser visível, aguarda reCAPTCHA (ou tenta automaticamente),
  salva sessão em session.json para não precisar mais de login.
- Execuções seguintes: carrega sessão salva, roda sem interação humana.
- Lê NCMs de bcoDados.xlsx (coluna A) e grava resultados nas colunas B+.
- Abas descobertas dinamicamente: qualquer aba nova encontrada no Econet gera
  colunas automaticamente no Excel.
"""

import asyncio
import io
import json
import os
import re
import shutil
import sys
from pathlib import Path

# Fix encoding no Windows (terminal cp1252 nao suporta Unicode)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

# ──────────────────────── Config ────────────────────────
BASE_DIR   = Path(__file__).parent
XLSX_PATH  = BASE_DIR / "bcoDados.xlsx"
SESSION    = BASE_DIR / "session.json"

ECONET_URL = "https://www.econeteditora.com.br/"
LOGIN      = os.getenv("ECONET_USERNAME", "")
SENHA      = os.getenv("ECONET_PASSWORD", "")

# Abas que nunca devem ser extraídas
SKIP_TABS = {"Exportação", "Importação", "Reforma Tributária", "Reforma Tributária - NOVO"}
# ────────────────────────────────────────────────────────


def formatar_ncm(raw) -> str:
    """84714190 → '8471.41.90'"""
    s = str(int(raw)).zfill(8)
    return f"{s[:4]}.{s[4:6]}.{s[6:]}"


def _abrir_workbook(read_only=True):
    """Abre bcoDados.xlsx, usando cópia temporária se estiver bloqueado."""
    try:
        wb = openpyxl.load_workbook(XLSX_PATH, read_only=read_only)
        return wb, None
    except PermissionError:
        tmp = XLSX_PATH.parent / "_tmp_read.xlsx"
        shutil.copy2(XLSX_PATH, tmp)
        return openpyxl.load_workbook(tmp, read_only=read_only), tmp


def ler_ncms(apenas_incompletos: bool = True) -> list[tuple[int, int]]:
    """Retorna lista de (linha_excel, ncm_int).

    apenas_incompletos=True  → somente linhas onde col B ou col D (PIS) está vazia.
    apenas_incompletos=False → todas as linhas com valor na col A.
    """
    wb, tmp = _abrir_workbook(read_only=True)
    try:
        ws = wb.active
        entradas = []
        for r in ws.iter_rows(min_row=2):
            if not r[0].value:
                continue
            if apenas_incompletos:
                incompleta = (len(r) < 2 or not r[1].value) or (len(r) < 4 or not r[3].value)
                if not incompleta:
                    continue
            ncm_int = int(str(r[0].value).replace(".", ""))
            entradas.append((r[0].row, ncm_int))
    finally:
        wb.close()
        if tmp and tmp.exists():
            try:
                tmp.unlink()
            except Exception:
                pass
    return entradas



async def fazer_login(page):
    print("🔐 Abrindo tela de login...")
    await page.goto(ECONET_URL, wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(3000)  # página estabilizar

    await page.wait_for_selector("text=Entrar", timeout=15000)
    await page.wait_for_timeout(1500)  # pausa antes de clicar
    await page.click("text=Entrar")
    await page.wait_for_selector("input[placeholder='Código / CPF']", state="visible", timeout=15000)
    await page.wait_for_timeout(1000)  # modal abriu, aguarda

    # digita usuário caractere a caractere como humano
    await page.type("input[placeholder='Código / CPF']", LOGIN, delay=120)
    await page.wait_for_timeout(700)
    await page.type("input[type='password']", SENHA, delay=100)
    await page.wait_for_timeout(1500)  # pausa antes do reCAPTCHA

    # ── Tenta resolver reCAPTCHA automaticamente ──
    print("🤖 Tentando resolver reCAPTCHA automaticamente...")
    resolved = await tentar_recaptcha(page)

    if not resolved:
        print("⚠️  reCAPTCHA precisa de interação humana.")
        print("   → Por favor, marque 'Não sou um robô' e clique em Entrar no browser.")
        await page.wait_for_selector("input[placeholder='Código / CPF']", state="hidden", timeout=120000)
        print("✅ Login detectado após reCAPTCHA manual.")
        return

    # Clica em Entrar (apenas se reCAPTCHA foi resolvido automaticamente)
    await page.wait_for_selector("#login_submit:not([disabled])", timeout=15000)
    await page.wait_for_timeout(1000)  # pausa natural antes de submeter
    await page.click("#login_submit")
    await page.wait_for_selector("input[placeholder='Código / CPF']", state="hidden", timeout=15000)
    print("✅ Login realizado com sucesso!")


async def tentar_recaptcha(page) -> bool:
    """
    Tenta clicar no checkbox do reCAPTCHA.
    Em browser real (não headless) geralmente passa automaticamente.
    """
    try:
        await page.wait_for_selector("iframe[src*='recaptcha']", timeout=8000)
        await page.wait_for_timeout(2000)  # iframe do reCAPTCHA carregar completamente

        recaptcha_frame = page.frame_locator("iframe[src*='recaptcha']").first
        await page.wait_for_timeout(1000)  # pausa antes de mover para o checkbox

        await recaptcha_frame.locator("#recaptcha-anchor").click(timeout=8000)
        await page.wait_for_timeout(4000)  # aguarda avaliação do Google

        checked = await recaptcha_frame.locator("#recaptcha-anchor").get_attribute("aria-checked")
        if checked == "true":
            print("✅ reCAPTCHA resolvido automaticamente!")
            return True

        # Se surgiu desafio de imagens, aguarda resolução
        await page.wait_for_timeout(6000)
        checked = await recaptcha_frame.locator("#recaptcha-anchor").get_attribute("aria-checked")
        return checked == "true"

    except Exception as e:
        print(f"   reCAPTCHA auto: {e}")
        return False


async def navegar_pis_cofins(page) -> str:
    """Navega até PIS/COFINS → Busca do Produto e retorna o src do iframe interno."""
    print("📂 Navegando para Federal → PIS/COFINS...")
    await page.wait_for_timeout(2000)

    # O menu usa CSS :hover — não funciona em headless. Navega o iframe #alvo
    # diretamente para a URL da seção PIS/COFINS (link tem target="alvo").
    pis_url = await page.evaluate("""() => {
        const a = document.querySelector('a[title="PIS / COFINS"]');
        return a ? a.href : '';
    }""")
    if not pis_url:
        pis_url = "https://www.econeteditora.com.br//pis_cofins/pis_cofins_capa.php"
    await page.evaluate(f"""() => {{
        const f = document.getElementById('alvo');
        if (f) f.src = '{pis_url}';
    }}""")
    await page.wait_for_timeout(4000)  # iframe #alvo carregar

    # Captura a URL de "Busca do Produto" no iframe #alvo e navega o iframe f2
    # diretamente — evita clicar em elemento que pode estar oculto
    busca_src = await page.evaluate("""() => {
        try {
            const f1 = document.getElementById('alvo');
            if (!f1 || !f1.contentDocument) return '';
            // Tenta pegar a URL do tab Busca do Produto
            const link = f1.contentDocument.querySelector('a[href*="form"]');
            if (link) {
                const href = link.href || link.getAttribute('href');
                // Normaliza para URL absoluta
                if (href.startsWith('http')) return href;
                return new URL(href, f1.src).href;
            }
            // Fallback: captura o iframe f2 já carregado
            const f2 = f1.contentDocument.querySelector('iframe');
            return f2 ? f2.src : '';
        } catch(e) { return ''; }
    }""")

    if busca_src and "form" in busca_src:
        # Navega f2 (container) para a URL de busca do produto
        await page.evaluate(f"""() => {{
            const f1 = document.getElementById('alvo');
            if (!f1 || !f1.contentDocument) return;
            const f2 = f1.contentDocument.querySelector('iframe');
            if (f2) f2.src = '{busca_src}';
        }}""")
        await page.wait_for_timeout(4000)  # f2 e f3 carregarem
    else:
        # Fallback: tenta clicar no tab
        busca_tab = page.frame_locator("#alvo").locator("text=Busca do Produto").first
        await busca_tab.click(timeout=15000)
        await page.wait_for_timeout(4000)

    # Captura URL do nível mais profundo (f3 se existir, senão f2)
    busca_src_real = await page.evaluate("""() => {
        const f1 = document.getElementById('alvo');
        if (!f1 || !f1.contentDocument) return '';
        const f2 = f1.contentDocument.querySelector('iframe');
        if (!f2) return '';
        if (f2.contentDocument) {
            const f3 = f2.contentDocument.querySelector('iframe');
            if (f3) return f3.src;
        }
        return f2.src;
    }""")

    print(f"✅ Seção PIS/COFINS aberta. iframe src: {busca_src_real[:80] if busca_src_real else '(vazio)'}...")
    return busca_src_real


async def _js_iframe(page: object, script: str):
    """Executa JS no iframe aninhado e retorna o resultado.
    Suporta 2 ou 3 níveis de iframe: alvo > f2 > f3 (se f3 existir, usa-o como f2).
    """
    return await page.evaluate(f"""() => {{
        try {{
            const f1 = document.getElementById('alvo');
            let f2 = f1.contentDocument.querySelector('iframe');
            // Econet adicionou nível extra: alvo > container > pis_cofins.php
            if (f2 && f2.contentDocument) {{
                const f3 = f2.contentDocument.querySelector('iframe');
                if (f3) f2 = f3;
            }}
            {script}
        }} catch(e) {{ return 'ERR: ' + e.message; }}
    }}""")


_JS_LINHAS_VISIVEIS = """
    const win = f2.contentDocument.defaultView || f2.contentWindow;
    function visivel(el) {
        try {
            let cur = el;
            while (cur && cur !== f2.contentDocument.body) {
                const st = win.getComputedStyle(cur);
                if (st.display === 'none' || st.visibility === 'hidden') return false;
                cur = cur.parentElement;
            }
            return true;
        } catch(e) { return true; }
    }
    const rows = Array.from(f2.contentDocument.querySelectorAll('table tr'))
        .filter(r => visivel(r));
    return rows.map(r =>
        Array.from(r.querySelectorAll('td,th'))
            .map(c => c.innerText.trim().replace(/\\n+/g,' '))
            .join('||')
    ).join('\\n');
"""


def _parse_observacoes(body_text: str) -> str:
    """Extrai conteúdo da seção Observações.
    Para antes de seções CST ou qualquer seção com 'Contribuintes'/'Contribuições'.
    """
    lines = body_text.splitlines()
    obs_start = -1
    for i, line in enumerate(lines):
        if line.strip() == "Observações":
            obs_start = i + 1
            break
    if obs_start == -1:
        return ""
    obs_lines = []
    for line in lines[obs_start:]:
        s = line.strip()
        if (
            "Código da Situação Tributária" in s
            or "Contribuintes" in s
            or "Contribuições" in s
        ):
            break
        if s:
            obs_lines.append(s)
    return "\n".join(obs_lines)


def _extrair_aba_completa(raw: str) -> dict:
    """Extrai PIS/COFINS separando Cumulativo e Não Cumulativo, ignorando Simples Nacional."""
    result = {"pis_cum": "", "cofins_cum": "", "pis_ncum": "", "cofins_ncum": "", "leg": ""}
    if not raw or raw.startswith("ERR"):
        return result

    in_aliquota = False
    legs = []
    for line in raw.splitlines():
        cells = [c.strip() for c in line.split("||")]
        if not any(cells):
            continue
        joined = " ".join(cells)
        if "Regime de Tributação" in joined or ("Alíquota" in joined and len(cells) == 1):
            in_aliquota = True
            continue
        if not in_aliquota:
            continue
        if len(cells) < 2:
            continue
        regime = cells[0]
        pis    = cells[1] if len(cells) > 1 else ""
        cofins = cells[2] if len(cells) > 2 else ""
        leg    = cells[3] if len(cells) > 3 else ""

        # Alíquota real é curta ("Vide observações"=16, "0,65%"=5) — texto longo é observação
        if len(pis) > 25:
            pis = ""
        if len(cofins) > 25:
            cofins = ""
        tem_valor = "%" in pis or "Vide" in pis or "%" in cofins or "Vide" in cofins
        if not tem_valor:
            continue
        if "Simples" in regime:
            continue  # ignora Simples Nacional
        if "Cumulativo" in regime and "Não" not in regime:
            result["pis_cum"]    = pis
            result["cofins_cum"] = cofins
            if leg:
                legs.append(f"Cum: {leg}")
        elif "Não Cumulativo" in regime:
            result["pis_ncum"]    = pis
            result["cofins_ncum"] = cofins
            if leg:
                legs.append(f"N.Cum: {leg}")
        elif regime:
            # Alíquota única (sem separação de regime)
            result["pis_cum"] = result["pis_ncum"] = pis
            result["cofins_cum"] = result["cofins_ncum"] = cofins
            if leg:
                legs.append(leg)

    result["leg"] = " | ".join(legs)
    return result


async def _clicar_aba(page, nome_aba: str) -> bool:
    """Clica na aba pelo nome no TabbedPanel do iframe. Retorna True se encontrou."""
    result = await _js_iframe(page, f"""
        const tabs = Array.from(f2.contentDocument.querySelectorAll('li.TabbedPanelsTab'));
        const tab = tabs.find(t => t.innerText.trim().includes('{nome_aba}'));
        if (tab) {{ tab.click(); return 'ok'; }}
        return 'not found';
    """)
    return result == "ok"


async def _listar_abas(page) -> list[str]:
    """Retorna lista com os nomes de todas as abas disponíveis no TabbedPanel do iframe."""
    result = await _js_iframe(page, """
        const tabs = Array.from(f2.contentDocument.querySelectorAll('li.TabbedPanelsTab'));
        return tabs.map(t => t.innerText.trim()).join('||');
    """)
    if not result or str(result).startswith("ERR"):
        return []
    return [t for t in result.split("||") if t]


async def buscar_ncm(page, ncm_formatado: str, busca_src: str) -> dict:
    """Busca um NCM via JS direto no iframe e retorna os dados extraídos."""
    print(f"  Buscando NCM {ncm_formatado}...")

    # Recarrega o iframe mais profundo (f3=pis_cofins.php, apelido f2 no _js_iframe)
    await page.evaluate(f"""() => {{
        const f1 = document.getElementById('alvo');
        const f2 = f1.contentDocument.querySelector('iframe');
        if (!f2) return;
        // Se há f3, recarrega f3; senão recarrega f2
        if (f2.contentDocument) {{
            const f3 = f2.contentDocument.querySelector('iframe');
            if (f3) {{ f3.src = '{busca_src}'; return; }}
        }}
        f2.src = '{busca_src}';
    }}""")
    await page.wait_for_timeout(3000)  # formulário recarregar

    # 1. Preenche campo NCM e submete
    await _js_iframe(page, f"""
        const inp = f2.contentDocument.getElementById('inpCodigoNcm');
        if (!inp) return 'ERR: inpCodigoNcm not found';
        inp.focus();
        inp.value = '{ncm_formatado}';
        return 'ok';
    """)
    await page.wait_for_timeout(1200)
    await _js_iframe(page, """
        const btn = f2.contentDocument.querySelector('input[value="Pesquisar"]');
        if (btn) btn.click();
        return 'ok';
    """)
    await page.wait_for_timeout(3500)

    # 2. Clica no radio do NCM mais específico sem "Ex "
    prefix = ncm_formatado[:7]
    await _js_iframe(page, f"""
        const rows = Array.from(f2.contentDocument.querySelectorAll('tr'));
        let clicked = false;
        for (const r of rows) {{
            const txt = r.innerText;
            if (txt.includes('{prefix}') && !txt.includes('Ex ')) {{
                const radio = r.querySelector('input[type="radio"]');
                if (radio) {{ radio.click(); clicked = true; break; }}
            }}
        }}
        if (!clicked) {{
            const first = f2.contentDocument.querySelector('input[type="radio"]');
            if (first) first.click();
        }}
        return 'ok';
    """)
    await page.wait_for_timeout(4000)  # página de detalhes do NCM carregar

    data = {
        "descricao": "", "pis_cum": "", "cofins_cum": "",
        "pis_ncum": "", "cofins_ncum": "", "regime": "", "legislacao": [],
        "obs_rg": "",
        "_abas": {},  # {"Nome da Aba": {"pis": ..., "cofins": ..., "leg": ..., "obs": ...}}
    }

    # ── 3. Regra Geral (aba padrão, já ativa) ──────────────────────────────
    raw = await _js_iframe(page, _JS_LINHAS_VISIVEIS)

    if not raw or raw.startswith("ERR"):
        print(f"     AVISO extração RG: {raw}")
    else:
        ncm_desc_rows = []
        in_aliquota = False
        for line in raw.splitlines():
            cells = [c.strip() for c in line.split("||")]
            if not any(cells):
                continue
            joined = " ".join(cells)
            if "Regime de Tributação" in joined or cells[0] == "Regime de Tributação":
                in_aliquota = True
                continue
            if "Alíquota" in joined and len(cells) == 1:
                in_aliquota = True
                continue
            if not in_aliquota:
                if len(cells) >= 2 and cells[0] not in ("", "NCM", "DESCRIÇÃO", "Produto"):
                    ncm_desc_rows.append(cells)
            else:
                if len(cells) < 3:
                    continue
                regime = cells[0]
                pis    = cells[1]
                cofins = cells[2]
                leg    = cells[3] if len(cells) > 3 else ""
                if not ("%" in pis or "Vide" in pis or "%" in cofins):
                    continue
                if "Cumulativo" in regime and "Não" not in regime:
                    data["pis_cum"]    = pis
                    data["cofins_cum"] = cofins
                    if leg: data["legislacao"].append(f"Cum: {leg}")
                    if not data["regime"]: data["regime"] = "Cumulativo / Não Cumulativo"
                elif "Não Cumulativo" in regime:
                    data["pis_ncum"]    = pis
                    data["cofins_ncum"] = cofins
                    if leg: data["legislacao"].append(f"N.Cum: {leg}")
                elif "Simples" in regime:
                    pass
                elif regime and "%" in pis:
                    data["pis_cum"] = data["pis_ncum"] = pis
                    data["cofins_cum"] = data["cofins_ncum"] = cofins
                    if leg: data["legislacao"].append(leg)
                    if not data["regime"]: data["regime"] = regime
        if ncm_desc_rows:
            last = ncm_desc_rows[-1]
            specific = last[1] if len(last) > 1 else last[0]

            # Busca a descrição da posição de 4 dígitos (ex: "3926") como base de contexto
            base = ""
            for row in ncm_desc_rows:
                code = str(row[0]).replace(".", "").strip()
                if code.isdigit() and len(code) == 4 and len(row) > 1:
                    base = row[1].strip()
                    break

            if base and base != specific.strip():
                data["descricao"] = f"{base} — {specific}"
            else:
                data["descricao"] = specific

    # Detecta regime especial e extrai Observações da Regra Geral
    body_rg = await _js_iframe(page, "return f2.contentDocument.body.innerText.substring(0,10000);")
    if isinstance(body_rg, str):
        if "Bebidas Frias" in body_rg:
            data["regime"] = "Bebidas Frias (Monofásico)"
            if not data["pis_cum"]:
                bf_raw = await _js_iframe(page, """
                    const win = f2.contentDocument.defaultView || f2.contentWindow;
                    function visivel(el) {
                        try {
                            let cur = el;
                            while (cur && cur !== f2.contentDocument.body) {
                                const st = win.getComputedStyle(cur);
                                if (st.display === 'none' || st.visibility === 'hidden') return false;
                                cur = cur.parentElement;
                            }
                            return true;
                        } catch(e) { return true; }
                    }
                    const rows = Array.from(f2.contentDocument.querySelectorAll('table tr'))
                        .filter(r => visivel(r));
                    for (const r of rows) {
                        const cells = Array.from(r.querySelectorAll('td,th'))
                            .map(c => c.innerText.trim().replace(/\\n+/g,' '));
                        if (cells.length >= 6 && cells[4].includes('%')) {
                            return cells[4] + '||' + cells[5];
                        }
                    }
                    return '';
                """)
                if bf_raw and "||" in bf_raw:
                    parts = bf_raw.split("||")
                    data["pis_cum"] = data["pis_ncum"] = parts[0].strip()
                    data["cofins_cum"] = data["cofins_ncum"] = parts[1].strip()
        elif "Incidência Monofásica" in body_rg or "Monofásico" in body_rg:
            data["regime"] = "Monofásico"
        data["obs_rg"] = _parse_observacoes(body_rg)

    # ── 4. Abas dinâmicas (todas exceto Regra Geral e SKIP_TABS) ──────────
    abas_disponiveis = await _listar_abas(page)
    for aba in abas_disponiveis:
        if aba == "Regra Geral" or aba in SKIP_TABS:
            continue
        if await _clicar_aba(page, aba):
            await page.wait_for_timeout(2000)
            raw_aba = await _js_iframe(page, _JS_LINHAS_VISIVEIS)
            body_aba = await _js_iframe(page, "return f2.contentDocument.body.innerText.substring(0,10000);")
            aba_data = _extrair_aba_completa(raw_aba)
            aba_data["obs"] = _parse_observacoes(body_aba) if isinstance(body_aba, str) else ""
            data["_abas"][aba] = aba_data
        else:
            print(f"     Aba '{aba}' não pôde ser clicada")

    data["legislacao"] = " | ".join(data["legislacao"])

    abas_extraidas = list(data["_abas"].keys())
    print(f"     OK PIS={data['pis_cum']} | COFINS={data['cofins_cum']} | Regime={data['regime']} | Abas={abas_extraidas}")
    return data


# Cabeçalhos fixos (cols 1–10) — sempre presentes
_FIXED_HEADERS = [
    "NCM", "NCM Econet", "Descrição",
    "PIS Cumulativo", "COFINS Cumulativo",
    "PIS Não Cumulativo", "COFINS Não Cumulativo",
    "Regime", "Legislação", "Observações (Regra Geral)",
]

# Larguras fixas para cols 1–10
_FIXED_WIDTHS = [12, 13, 55, 18, 20, 20, 22, 28, 65, 55]


def _registrar_historico(ws_hist, ncm_fmt: str, dados_ant: dict, dados_nov: dict) -> int:
    """Grava no sheet Histórico o registro inicial ou mudanças detectadas.

    dados_ant: {header_name: old_value} vindo do snapshot do Excel
    dados_nov: {header_name: new_value} com os cabeçalhos como chave
    Retorna o número de linhas inseridas.
    """
    from datetime import datetime
    agora = datetime.now().strftime("%Y-%m-%d %H:%M")

    is_novo = not any(
        str(dados_ant.get(h) or "").strip()
        for h in ["PIS Cumulativo", "COFINS Cumulativo", "Regime"]
    )

    linhas = 0
    for header, nov in dados_nov.items():
        ant = str(dados_ant.get(header) or "").strip()
        nov_str = str(nov or "").strip()

        if is_novo:
            if nov_str:
                ws_hist.append([agora, ncm_fmt, "Registro Inicial", header, "—", nov_str])
                linhas += 1
        else:
            if ant != nov_str:
                ws_hist.append([agora, ncm_fmt, "Atualização", header, ant, nov_str])
                linhas += 1
    return linhas


def salvar_excel(entradas: list[tuple[int, int]], resultados: list[dict]):
    destino = XLSX_PATH
    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
    except PermissionError:
        destino = XLSX_PATH.parent / "bcoDados_resultado.xlsx"
        print(f"⚠️  Excel aberto — salvando em {destino.name}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plan1"

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF")

    # Garante cabeçalhos fixos nas cols 1–10
    for ci, h in enumerate(_FIXED_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Constrói mapa de colunas a partir dos cabeçalhos existentes (cols 11+)
    col_map: dict[str, int] = {h: i for i, h in enumerate(_FIXED_HEADERS, 1)}
    for cell in ws[1]:
        if cell.value and cell.value not in col_map:
            col_map[cell.value] = cell.column

    # Descobre todas as abas nos resultados e cria colunas se necessário
    abas_vistas: list[str] = []
    seen: set[str] = set()
    for r in resultados:
        for aba in r.get("_abas", {}):
            if aba not in seen:
                abas_vistas.append(aba)
                seen.add(aba)

    for aba in abas_vistas:
        sub_headers = [
            f"{aba} - Legislação", f"{aba} - Observações",
            f"{aba} - PIS Cumulativo", f"{aba} - COFINS Cumulativo",
            f"{aba} - PIS Não Cumulativo", f"{aba} - COFINS Não Cumulativo",
        ]
        for sh in sub_headers:
            if sh not in col_map:
                next_col = max(col_map.values()) + 1
                col_map[sh] = next_col
                c = ws.cell(row=1, column=next_col, value=sh)
                c.fill = hfill
                c.font = hfont
                c.alignment = Alignment(horizontal="center", wrap_text=True)
                col_letter = openpyxl.utils.get_column_letter(next_col)
                if "PIS" in sh or "COFINS" in sh:
                    ws.column_dimensions[col_letter].width = 16
                elif "Legislação" in sh:
                    ws.column_dimensions[col_letter].width = 50
                else:
                    ws.column_dimensions[col_letter].width = 55
                print(f"     ➕ Nova coluna criada: {sh} (col {next_col})")

    # Snapshot do estado atual para comparação no Histórico
    snapshot: dict[int, dict] = {}
    for r in ws.iter_rows(min_row=2):
        if r[0].value:
            row_data = {}
            for header, col in col_map.items():
                idx = col - 1
                row_data[header] = r[idx].value if idx < len(r) else None
            snapshot[r[0].row] = row_data

    # Sheet Histórico — cria ou abre, garantindo schema com coluna Tipo
    hist_headers = ["Data/Hora", "NCM", "Tipo", "Campo", "Valor Anterior", "Valor Novo"]
    if "Histórico" not in wb.sheetnames:
        ws_hist = wb.create_sheet("Histórico")
        ws_hist.append(hist_headers)
        hfill_h = PatternFill("solid", fgColor="1F4E79")
        hfont_h = Font(bold=True, color="FFFFFF")
        for ci, h in enumerate(hist_headers, 1):
            c = ws_hist.cell(row=1, column=ci, value=h)
            c.fill = hfill_h
            c.font = hfont_h
            c.alignment = Alignment(horizontal="center")
        for w, col in zip([18, 13, 16, 22, 40, 40], ["A", "B", "C", "D", "E", "F"]):
            ws_hist.column_dimensions[col].width = w
    else:
        ws_hist = wb["Histórico"]
        # Migra header antigo (5 colunas) para novo (6 colunas com Tipo)
        if ws_hist.cell(1, 3).value != "Tipo":
            ws_hist.insert_cols(3)
            ws_hist.cell(1, 3, "Tipo")
            hfill_h = PatternFill("solid", fgColor="1F4E79")
            hfont_h = Font(bold=True, color="FFFFFF")
            for ci in range(1, 7):
                ws_hist.cell(1, ci).fill = hfill_h
                ws_hist.cell(1, ci).font = hfont_h
                ws_hist.cell(1, ci).alignment = Alignment(horizontal="center")
            for w, col in zip([18, 13, 16, 22, 40, 40], ["A", "B", "C", "D", "E", "F"]):
                ws_hist.column_dimensions[col].width = w

    total_linhas_hist = 0
    total_cols = max(col_map.values())

    for (ri, ncm), r in zip(entradas, resultados):
        dados_ant = snapshot.get(ri, {})

        # Monta dict de novos dados com chave = cabeçalho do Excel
        dados_nov = {
            "Descrição":                r["descricao"],
            "PIS Cumulativo":           r["pis_cum"],
            "COFINS Cumulativo":        r["cofins_cum"],
            "PIS Não Cumulativo":       r["pis_ncum"],
            "COFINS Não Cumulativo":    r["cofins_ncum"],
            "Regime":                   r["regime"],
            "Legislação":               r["legislacao"],
            "Observações (Regra Geral)": r["obs_rg"],
        }
        for aba, aba_data in r.get("_abas", {}).items():
            dados_nov[f"{aba} - PIS Cumulativo"]        = aba_data.get("pis_cum", "")
            dados_nov[f"{aba} - COFINS Cumulativo"]     = aba_data.get("cofins_cum", "")
            dados_nov[f"{aba} - PIS Não Cumulativo"]    = aba_data.get("pis_ncum", "")
            dados_nov[f"{aba} - COFINS Não Cumulativo"] = aba_data.get("cofins_ncum", "")
            dados_nov[f"{aba} - Legislação"]            = aba_data.get("leg", "")
            dados_nov[f"{aba} - Observações"]           = aba_data.get("obs", "")

        total_linhas_hist += _registrar_historico(ws_hist, formatar_ncm(ncm), dados_ant, dados_nov)

        # Grava NCM (col 1) e NCM Econet (col 2) fixos
        ws.cell(ri, 1, ncm)
        ws.cell(ri, 2, formatar_ncm(ncm))

        # Grava demais campos pelo col_map
        for header, value in dados_nov.items():
            col = col_map.get(header)
            if col:
                ws.cell(ri, col, value)

        for ci in range(1, total_cols + 1):
            ws.cell(ri, ci).alignment = Alignment(wrap_text=True, vertical="top")

    # Larguras para colunas fixas 1–10
    for ci, w in enumerate(_FIXED_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 30
    wb.save(destino)
    print(f"\n💾 Salvo: {destino}")
    if total_linhas_hist:
        print(f"📝 {total_linhas_hist} linha(s) adicionada(s) ao Histórico.")
    else:
        print("✅ Nenhuma mudança detectada — Histórico não alterado.")


async def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--todos", action="store_true", help="Varre todos os NCMs, incluindo já preenchidos")
    parser.add_argument("--ncms", type=str, default="", help="Lista de NCMs separados por vírgula para varredura seletiva")
    args = parser.parse_args()

    if args.ncms:
        # Modo seletivo: apenas os NCMs informados via --ncms
        ncms_filtro = set(c.strip().replace(".", "") for c in args.ncms.split(",") if c.strip())
        todas = ler_ncms(apenas_incompletos=False)
        entradas = [(row, ncm) for row, ncm in todas if str(ncm) in ncms_filtro]
        total = len(entradas)
        print(f"\n{'='*55}")
        print(f"🎯 MODO SELETIVO — {total} NCM(s) na fila")
        print(f"{'='*55}")
        for i, (_, ncm) in enumerate(entradas, 1):
            print(f"   {i:>2}. {formatar_ncm(ncm)}  ({ncm})")
        print(f"{'='*55}\n")
    else:
        modo_todos = args.todos
        entradas = ler_ncms(apenas_incompletos=not modo_todos)
        total = len(entradas)
        ncms = [ncm for _, ncm in entradas]
        if modo_todos:
            print(f"\n{'='*55}")
            print(f"🔄 MODO TODOS — {total} NCM(s) na fila (incluindo já preenchidos)")
            print(f"{'='*55}")
            for i, n in enumerate(ncms, 1):
                print(f"   {i:>2}. {formatar_ncm(n)}  ({n})")
            print(f"{'='*55}\n")
        else:
            print(f"\n{'='*55}")
            print(f"📋 MODO PENDENTES — {total} NCM(s) sem dados encontrados")
            print(f"{'='*55}")
            for i, n in enumerate(ncms, 1):
                print(f"   {i:>2}. {formatar_ncm(n)}  ({n})")
            print(f"{'='*55}\n")

    if not entradas:
        print("✅ Nenhum NCM para processar. Encerrando.")
        return

    sessao_existe = SESSION.exists()

    async with async_playwright() as pw:
        # Usa Chromium visível na 1ª execução, headless depois
        browser = await pw.chromium.launch(
            headless=sessao_existe,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox", "--disable-gpu"]
        )

        ctx_args = {"storage_state": str(SESSION)} if sessao_existe else {}
        context = await browser.new_context(
            **ctx_args,
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1400, "height": 900}
        )
        page = await context.new_page()

        # Login se não há sessão
        if not sessao_existe:
            await fazer_login(page)
            await context.storage_state(path=str(SESSION))
            print(f"💾 Sessão salva em {SESSION} — próximas execuções sem login!")
        else:
            print("✅ Sessão carregada — sem necessidade de login/reCAPTCHA!")
            await page.goto(ECONET_URL, wait_until="domcontentloaded", timeout=60000)
            # Verifica se sessão ainda é válida
            try:
                await page.wait_for_selector("text=Minha Biblioteca", timeout=5000)
            except PlaywrightTimeout:
                print("⚠️  Sessão expirada. Fazendo novo login...")
                SESSION.unlink(missing_ok=True)
                await fazer_login(page)
                await context.storage_state(path=str(SESSION))

        busca_src = await navegar_pis_cofins(page)

        total = len(entradas)
        resultados = []
        for idx, (_, ncm) in enumerate(entradas, 1):
            ncm_fmt = formatar_ncm(ncm)
            print(f"\n{'─'*55}")
            print(f"▶  Varrendo {idx}/{total}: {ncm_fmt}  ({ncm})")
            ncms_restantes = [formatar_ncm(n) for _, n in entradas[idx:]]
            if ncms_restantes:
                print(f"   Restantes: {', '.join(ncms_restantes)}")
            else:
                print(f"   Restantes: (nenhum — este é o último)")
            print(f"{'─'*55}")
            try:
                dados = await buscar_ncm(page, ncm_fmt, busca_src)
            except Exception as e:
                print(f"  ❌ ERRO no NCM {ncm_fmt}: {e}")
                dados = {
                    "descricao": "", "pis_cum": "", "cofins_cum": "",
                    "pis_ncum": "", "cofins_ncum": "", "regime": "",
                    "legislacao": "", "obs_rg": "", "_abas": {},
                }
            resultados.append(dados)

        await browser.close()

    salvar_excel(entradas, resultados)
    print("\n🎉 Concluído! Todos os NCMs processados.")


if __name__ == "__main__":
    asyncio.run(main())
