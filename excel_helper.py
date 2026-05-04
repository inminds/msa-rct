"""
excel_helper.py — CLI helper for Node.js to read/write bcoDados.xlsx

Usage:
  python excel_helper.py read                       → JSON array of all rows
  python excel_helper.py add NCM1 ...               → add NCMs to col A if not present
  python excel_helper.py restore NCM CAMPO VALOR    → restore old value for a field
"""
import sys
import json
import shutil
import tempfile
import os

EXCEL_PATH = "bcoDados.xlsx"
SHEET_NAME = "Plan1"

# Fixed headers that occupy columns 1-10
FIXED_HEADERS = [
    "NCM", "NCM Econet", "Descrição",
    "PIS Cumulativo", "COFINS Cumulativo",
    "PIS Não Cumulativo", "COFINS Não Cumulativo",
    "Regime", "Legislação", "Observações (Regra Geral)",
]


def _open_wb(path):
    try:
        import openpyxl
        return openpyxl.load_workbook(path)
    except PermissionError:
        tmp = shutil.copy2(path, tempfile.mktemp(suffix=".xlsx"))
        import openpyxl
        return openpyxl.load_workbook(tmp)


def cmd_read():
    if not os.path.exists(EXCEL_PATH):
        print(json.dumps([]))
        return

    wb = _open_wb(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    # Build col_map from header row
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.column] = str(cell.value).strip()

    ncm_col = None
    for col, h in headers.items():
        if h == "NCM":
            ncm_col = col
            break
    if ncm_col is None:
        ncm_col = 1

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ncm_val = row[ncm_col - 1] if len(row) >= ncm_col else None
        if not ncm_val:
            continue
        entry = {}
        for col_idx, h in headers.items():
            val = row[col_idx - 1] if len(row) >= col_idx else None
            entry[h] = str(val).strip() if val is not None else ""
        rows.append(entry)

    print(json.dumps(rows, ensure_ascii=True))


def cmd_add(ncm_codes: list[str]):
    if not os.path.exists(EXCEL_PATH):
        # Create fresh workbook
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        for i, h in enumerate(FIXED_HEADERS, 1):
            ws.cell(row=1, column=i, value=h)
        existing = set()
    else:
        wb = _open_wb(EXCEL_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

        # Ensure header row has NCM in col 1
        if ws.cell(row=1, column=1).value != "NCM":
            ws.cell(row=1, column=1, value="NCM")

        # Collect existing NCMs
        existing = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                existing.add(str(row[0]).strip())

    added = []
    for ncm in ncm_codes:
        ncm = str(ncm).strip()
        if ncm and ncm not in existing:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=ncm)
            existing.add(ncm)
            added.append(ncm)

    try:
        wb.save(EXCEL_PATH)
    except PermissionError:
        fallback = "bcoDados_resultado.xlsx"
        wb.save(fallback)
        print(json.dumps({"added": added, "saved_to": fallback}, ensure_ascii=True))
        return

    print(json.dumps({"added": added, "saved_to": EXCEL_PATH}, ensure_ascii=True))


def cmd_restore(ncm: str, field: str, value: str):
    """Restore (overwrite) a single field for a given NCM in the Excel."""
    # Map field name → column index (1-based)
    FIELD_COL = {
        "Descrição": 3,             # C
        "PIS Cumulativo": 4,        # D
        "COFINS Cumulativo": 5,     # E
        "PIS Não Cumulativo": 6,    # F
        "COFINS Não Cumulativo": 7, # G
        "Regime": 8,                # H
    }

    col_idx = FIELD_COL.get(field)
    if col_idx is None:
        print(json.dumps({"error": f"Campo desconhecido: {field}"}, ensure_ascii=True))
        sys.exit(1)

    if not os.path.exists(EXCEL_PATH):
        print(json.dumps({"error": "Excel não encontrado"}, ensure_ascii=True))
        sys.exit(1)

    import openpyxl
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except PermissionError:
        print(json.dumps({"error": "Excel está aberto — feche e tente novamente"}, ensure_ascii=True))
        sys.exit(1)

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    # Find the row for this NCM
    target_row = None
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip() == ncm.strip():
            target_row = cell.row
            break

    if target_row is None:
        print(json.dumps({"error": f"NCM {ncm} não encontrado no Excel"}, ensure_ascii=True))
        sys.exit(1)

    ws.cell(row=target_row, column=col_idx, value=value if value else None)

    try:
        wb.save(EXCEL_PATH)
    except PermissionError:
        fallback = "bcoDados_resultado.xlsx"
        wb.save(fallback)
        print(json.dumps({"restored": True, "saved_to": fallback}, ensure_ascii=True))
        return

    print(json.dumps({"restored": True, "ncm": ncm, "field": field, "value": value, "saved_to": EXCEL_PATH}, ensure_ascii=True))


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python excel_helper.py read | add NCM1 NCM2 ... | restore NCM CAMPO VALOR", file=sys.stderr)
        sys.exit(1)

    cmd = sys.argv[1].lower()
    if cmd == "read":
        cmd_read()
    elif cmd == "add":
        cmd_add(sys.argv[2:])
    elif cmd == "restore":
        if len(sys.argv) < 5:
            print("Usage: python excel_helper.py restore NCM CAMPO VALOR", file=sys.stderr)
            sys.exit(1)
        cmd_restore(sys.argv[2], sys.argv[3], sys.argv[4])
    else:
        print(f"Unknown command: {cmd}", file=sys.stderr)
        sys.exit(1)
